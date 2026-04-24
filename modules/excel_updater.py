"""
excel_updater.py — V03 format
Flow mỗi ngày:
  1. Load 4 RAW files mới → update 4 RAW sheets trong report
  2. Tính toán từ RAW → upsert row vào SUMMARY (1 row/ngày)
  3. Rebuild WEEKLY + MONTHLY từ toàn bộ SUMMARY
  4. Update PR RW LOT LIST (dedup by Lot+HoldDate+HoldCode)
  5. Save YYYYMMDD_IPSS_DAILY_REPORT_V03.xlsx
"""
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from collections import defaultdict
import os, shutil, glob

# ── Import từ modules cùng package ──────────────────────────────────────────
from modules.calculator import (
    STEPS, STEP_NAMES, DI_RINSE_STEPS,
    calc_wip_by_step, calc_movement_by_step, calc_develop_move,
    calc_input, calc_pr_rw, get_week_label, upsert_trend
)

# PR codes V03 (sorted PR01-PR33 + PR99)
PR_COLS = [f"PR{i:02d}" for i in range(1, 34)] + ["PR99"]
PR_DESC = {
    "PR01":"Exposure NG","PR02":"Coating NG-SpeedBoat","PR03":"Developer NG",
    "PR04":"Operator Mistake","PR05":"Wafer Rotation NG","PR06":"MIS-Align",
    "PR07":"Align Key Damage","PR08":"Particle","PR09":"Scratch","PR10":"FOCUS",
    "PR11":"WET NG","PR12":"Equipment NG","PR13":"Reticle Defect","PR14":"He Error",
    "PR15":"PR RW Electric off","PR16":"Coating NG-Discolor","PR17":"Coating NG-Bubble",
    "PR18":"DH Over","PR19":"Eng'r Check Machine","PR20":"Cu Leak",
    "PR21":"Reticle Rotation","PR22":"6PRM Overhang","PR23":"PR Remain",
    "PR24":"Eng'r Depo Req","PR25":"Eng'r Wet Req","PR26":"Eng'r Etch Req",
    "PR27":"Eng'r Photo Req","PR28":"Eng'r PM Req","PR29":"Substrate Warp",
    "PR30":"U-GAN HCL RW","PR31":"Eng RND Req","PR32":"SiO2 Defect",
    "PR33":"Pre-Test EQ","PR99":"Etc",
}
DEVELOP_STEP = "E3157"  # DF Rate = PR RW / E3157 × 100

# Styles
THIN = Side(style="thin",  color="D0D0D0")
MED  = Side(style="medium",color="9FA8B4")
BT   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
C_NAVY="1F3864"; C_BLUE="2F5597"; C_TEAL="00695C"; C_PURP="4A235A"
DATE_FMT="yyyy-mm-dd"; DT_FMT="dd/mm/yyyy hh:mm"; PCT_FMT='0.00"%"'

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(s=10, b=False, c="000000"): return Font(name="Arial", size=s, bold=b, color=c)

def _copy_row_format(ws, src_row, tgt_row, max_col=100):
    """Copies format and formulas from source row to target row using official Translator."""
    from openpyxl.formula.translate import Translator
    for c in range(1, max_col + 1):
        src_cell = ws.cell(src_row, c)
        tgt_cell = ws.cell(tgt_row, c)
        
        # 1. Copy Value/Formula using Translator (Handles relative/absolute refs correctly)
        if src_cell.data_type == 'f':
            formula = src_cell.value
            # Translate from src_cell coordinate (e.g. 'A107') to tgt_cell (e.g. 'A108')
            new_formula = Translator(formula, origin=src_cell.coordinate).translate_formula(tgt_cell.coordinate)
            tgt_cell.value = new_formula
        
        # 2. Copy Style
        if src_cell.has_style:

            tgt_cell.font = openpyxl.styles.Font(
                name=src_cell.font.name, size=src_cell.font.size,
                bold=src_cell.font.bold, italic=src_cell.font.italic, color=src_cell.font.color
            )
            tgt_cell.border = openpyxl.styles.Border(
                left=src_cell.border.left, right=src_cell.border.right,
                top=src_cell.border.top, bottom=src_cell.border.bottom
            )
            tgt_cell.fill = openpyxl.styles.PatternFill(
                fill_type=src_cell.fill.fill_type, fgColor=src_cell.fill.fgColor
            )
            tgt_cell.alignment = openpyxl.styles.Alignment(
                horizontal=src_cell.alignment.horizontal, vertical=src_cell.alignment.vertical
            )
            tgt_cell.number_format = src_cell.number_format


def _font(b=False, c="000000", s=10): return Font(bold=b, color=c, size=s, name="Arial")
def _align(h="center",v="center",w=False): return Alignment(horizontal=h,vertical=v,wrap_text=w)
def _si(v):
    if pd.isna(v) or v == "": return 0
    try:
        s = str(v).replace(",","").replace(" ","").replace("\xa0","").strip()
        if not s: return 0
        return int(float(s))
    except:
        return 0
def _norm(s): return pd.to_datetime(s, errors="coerce")
def _date(v):
    if v is None or (isinstance(v,float) and np.isnan(v)): return None
    try: return pd.Timestamp(v).date()
    except: return None

def _dfr(prt, dev):
    """DF Rate as percent (e.g. 74.31), NOT decimal"""
    return round(prt/dev*100, 2) if dev and dev > 0 else 0.0

def _hc(ws, r, c, v, bg=C_NAVY, fg="FFFFFF", b=True, s=10, wrap=False):
    cl=ws.cell(r,c,v); cl.fill=_fill(bg); cl.font=_font(b,fg,s)
    cl.alignment=_align("center","center",wrap); cl.border=BT; return cl

def _dc(ws, r, c, v, bg="FFFFFF", b=False, h="right", fmt=None):
    cl=ws.cell(r,c,v); cl.fill=_fill(bg); cl.font=_font(b)
    cl.alignment=_align(h,"center"); cl.border=BT
    if fmt: cl.number_format=fmt
    elif isinstance(v, (int, float)): cl.number_format = '#,##0'
    elif isinstance(v, (datetime, pd.Timestamp)): cl.number_format = DATE_FMT
    return cl


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1: UPDATE 4 RAW SHEETS
# ══════════════════════════════════════════════════════════════════════════════

def _write_raw_sheet(wb, sheet_name: str, df_new: pd.DataFrame):
    """Replace RAW sheet content. We clear and rewrite to preserve links."""
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
        # Clear existing data
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
    
    ws.sheet_view.showGridLines = False
    
    # ALWAYS clear and write headers, even if df_new is empty (Ghi nối tiếp nhưng RAW phải tươi)
    cols = list(df_new.columns) if not df_new.empty else ["DATA NOT FOUND FOR TODAY"]
    for ci, col in enumerate(cols, 1):
        _hc(ws, 1, ci, str(col), bg=C_NAVY, s=9)

    if df_new.empty:
        # If no new data, we leave it empty (only headers) to avoid showing old data
        return ws

    for ri, (_, row) in enumerate(df_new.iterrows(), 2):
        pg = str(row.get("Product Group", "")).upper() if "Product Group" in df_new.columns else ""
        bg = "EBF5FB" if "IPSS" in pg else ("FFFFFF" if ri % 2 == 0 else "FAFAFA")
        for ci, col in enumerate(cols, 1):
            v = row[col]
            if isinstance(v, pd.Timestamp): v = _date(v)
            elif pd.isna(v) if not isinstance(v, str) else False: v = None
            cl = ws.cell(ri, ci, v)
            cl.fill = _fill(bg); cl.font = _font(s=9)
            cl.alignment = _align("left" if isinstance(v,str) else "center","center")
            cl.border = BT
            if isinstance(v, __import__('datetime').date): cl.number_format = DATE_FMT

    for ci in range(1, len(cols)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2: CALCULATE DAILY ROW FROM RAW
# ══════════════════════════════════════════════════════════════════════════════

def calc_daily_row_from_raw(raw_data: dict, today_dt: datetime, week_start_day: int = 4) -> dict:
    """
    Tính toán 1 row SUMMARY từ 4 RAW sheets.
    Logic:
      - WIP: từ RAW_WIP snapshot (IPSS only)
      - Move: từ RAW_MOVE TrackOut (IPSS only, hôm nay)
      - Input: từ RAW_INPUT E1100 IPSS hôm nay
      - PR RW: từ RAW_HOLD HISTORY PR codes (IPSS, ngày hôm nay)
      - E3157: Move step E3157 hôm nay
      - DF Rate = PR RW qty / E3157 * 100
    """
    df_wip   = raw_data.get("wip",   pd.DataFrame())
    df_move  = raw_data.get("move",  pd.DataFrame())
    df_input = raw_data.get("input", pd.DataFrame())
    df_hold  = raw_data.get("hold_history", pd.DataFrame())

    # Normalize dates
    for df, col in [(df_move,"Date"),(df_input,"Date")]:
        if not df.empty and col in df.columns:
            df[col] = _norm(df[col])
    hold_col = next((c for c in df_hold.columns if "Hold date" in str(c)), None) if not df_hold.empty else None
    if hold_col and not df_hold.empty:
        df_hold[hold_col] = _norm(df_hold[hold_col])

    # Filter IPSS only
    def ipss_filter(df):
        if "Product Group" in df.columns:
            return df[df["Product Group"].astype(str).str.upper().str.contains("IPSS", na=False)]
        return df

    df_move_ipss  = ipss_filter(df_move)
    df_wip_ipss   = ipss_filter(df_wip)
    df_input_ipss = ipss_filter(df_input)
    df_input_pss  = df_input[~df_input.index.isin(df_input_ipss.index)] if not df_input.empty else pd.DataFrame()

    def get_work_date(dt):
        if pd.isna(dt): return None
        return (dt + pd.Timedelta(hours=4)).floor('D')

    # WIP snapshot (point-in-time, all IPSS lots)
    wip = {s: 0 for s in STEPS}
    if not df_wip_ipss.empty and "Step" in df_wip_ipss.columns:
        for step in STEPS:
            if step == "E1500":
                # E1500 = tổng DI Rinse mới (E1500) + DI Rinse cũ (E3100)
                qty = sum(
                    _si(df_wip_ipss[df_wip_ipss["Step"].astype(str).str.strip() == s]["Qty"].sum())
                    if "Qty" in df_wip_ipss.columns else 0
                    for s in DI_RINSE_STEPS
                )
                wip[step] = qty
            else:
                m = df_wip_ipss[df_wip_ipss["Step"].astype(str).str.strip() == step]
                wip[step] = _si(m["Qty"].sum()) if "Qty" in m.columns else 0

    # Movement (TrackOut IPSS) - use shift logic
    move_today = {s: 0 for s in STEPS}
    if not df_move_ipss.empty:
        target_norm = pd.Timestamp(today_dt).normalize()
        # Đảm bảo cột Date là datetime
        date_col = next((c for c in df_move_ipss.columns if "Date" in str(c) or "Time" in str(c)), "Date")
        if date_col in df_move_ipss.columns:
            df_move_ipss[date_col] = pd.to_datetime(df_move_ipss[date_col], errors="coerce")
            df_move_ipss["WorkDate"] = df_move_ipss[date_col].apply(get_work_date)
            df_today_move = df_move_ipss[df_move_ipss["WorkDate"] == target_norm]
            
            if "EventName" in df_today_move.columns:
                df_today_move = df_today_move[df_today_move["EventName"].astype(str).str.contains("TrackOut", na=False)]
            elif "Event Name" in df_today_move.columns:
                df_today_move = df_today_move[df_today_move["Event Name"].astype(str).str.contains("TrackOut", na=False)]

            for step in STEPS:
                if step == "E1500":
                    # E1500 = tổng DI Rinse mới (E1500) + DI Rinse cũ (E3100)
                    qty = 0
                    if "Step" in df_today_move.columns and "Qty" in df_today_move.columns:
                        for s in DI_RINSE_STEPS:
                            qty += _si(df_today_move[df_today_move["Step"].astype(str).str.strip() == s]["Qty"].sum())
                    move_today[step] = qty
                else:
                    m = df_today_move[df_today_move["Step"].astype(str).str.strip() == step] if "Step" in df_today_move.columns else pd.DataFrame()
                    move_today[step] = _si(m["Qty"].sum()) if "Qty" in m.columns else 0

    # Input IPSS (E1100) - use shift logic
    input_ipss_qty = 0
    input_pss_qty  = 0
    if not df_input.empty:
        target_norm = pd.Timestamp(today_dt).normalize()
        date_col = next((c for c in df_input.columns if "Date" in str(c) or "Time" in str(c)), "Date")
        if date_col in df_input.columns:
            df_input[date_col] = pd.to_datetime(df_input[date_col], errors="coerce")
            df_input["WorkDate"] = df_input[date_col].apply(get_work_date)
            df_inp_day = df_input[df_input["WorkDate"] == target_norm].copy()
            
            if "Step" in df_inp_day.columns:
                df_inp_day = df_inp_day[df_inp_day["Step"].astype(str).str.strip() == "E1100"]
                
            qty_col = next((c for c in df_inp_day.columns if "Qty" in str(c)), "Qty")
            if qty_col in df_inp_day.columns:
                df_inp_day[qty_col] = pd.to_numeric(df_inp_day[qty_col], errors="coerce").fillna(0)
                for _, r in df_inp_day.iterrows():
                    pg = str(r.get("Product Group", r.get("ProductGroup", ""))).upper()
                    qty = _si(r.get(qty_col, 0))
                    if "IPSS" in pg: input_ipss_qty += qty
                    else: input_pss_qty += qty

    # PR RW (IPSS, PR codes) - use shift logic
    pr_by_code = defaultdict(int)
    pr_rw_total = 0
    if not df_hold.empty and hold_col:
        target_norm = pd.Timestamp(today_dt).normalize()
        df_hold[hold_col] = pd.to_datetime(df_hold[hold_col], errors="coerce")
        df_hold["WorkDate"] = df_hold[hold_col].apply(get_work_date)
        df_hold_day = df_hold[df_hold["WorkDate"] == target_norm]
        
        # Filter for IPSS - be slightly flexible for PR rework to avoid 'missing' lots
        if "Product Group" in df_hold_day.columns:
            # Keep if IPSS is in group OR if it's a PR code and group is empty/N/A
            pg_mask = df_hold_day["Product Group"].astype(str).str.upper().str.contains("IPSS", na=False)
            if "Hold Code" in df_hold_day.columns:
                is_pr = df_hold_day["Hold Code"].astype(str).str.upper().str.startswith("PR")
                pg_mask |= (is_pr & df_hold_day["Product Group"].isna())
            df_hold_day = df_hold_day[pg_mask]
        
        if "Hold Code" in df_hold_day.columns and "Lot" in df_hold_day.columns:
            pr_df = df_hold_day[df_hold_day["Hold Code"].astype(str).str.upper().str.startswith("PR")].copy()
            # Dedup to match Lot List logic (Lot + Code)
            pr_df = pr_df.sort_values(["Lot", "Hold Code"], ascending=[True, True])
            pr_df = pr_df.drop_duplicates(subset=["Lot", "Hold Code"], keep="first")
            
            for _, r in pr_df.iterrows():
                code = str(r.get("Hold Code","")).strip()
                qty  = _si(r.get("Qty",0))
                if code in PR_COLS:
                    pr_by_code[code] += qty
                    pr_rw_total += qty
    
    e3157_move = move_today.get(DEVELOP_STEP, 0)
    df_rate    = _dfr(pr_rw_total, e3157_move)
    total_wip  = sum(wip.values())
    total_move = sum(move_today.values())

    row = {
        "Date":          today_dt.strftime("%Y-%m-%d"),
        "Label":         today_dt.strftime("%Y-%m-%d"),
        "Year":          today_dt.year,
        "Month":         today_dt.month,
        "Week":          today_dt.isocalendar()[1], 
        "Week_ISO":      today_dt.isocalendar()[1], 
        "MONTH_Label":   today_dt.strftime("%Y-%m"),
        "Week_Label":    get_week_label(today_dt, week_start_day),
        "Input_PSS":     input_pss_qty,
        "Input_IPSS":    input_ipss_qty,
        "EPI_Shipment":  0, 
        "Total_WIP":     total_wip,
        "Total_Move":    total_move,
        "E3157_Move":    e3157_move,
        "PR_RW_Qty":     pr_rw_total,
        "DF_Rate":       df_rate,
    }
    for s in STEPS:
        row[f"WIP_{s}"]  = wip.get(s, 0)
        row[f"Move_{s}"] = move_today.get(s, 0)
    for code in PR_COLS:
        row[f"PR_{code}"] = pr_by_code.get(code, 0)

    return row


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3: READ/WRITE SUMMARY SHEET
# ══════════════════════════════════════════════════════════════════════════════

SUMMARY_LABEL_COL = "Ngày"  # col A label in SUMMARY

def read_summary(wb_instance=None, file_path=None) -> pd.DataFrame:
    """Read existing SUMMARY sheet from Excel file or workbook object."""
    # Priority 1: Use pandas on file_path (Most robust)
    if file_path and os.path.exists(file_path):
        try:
            xl = pd.ExcelFile(file_path, engine="openpyxl")
            target_sheet = next((s for s in xl.sheet_names if s.strip().upper() == "SUMMARY"), None)
            if target_sheet:
                # Find header row by scanning first 15 rows
                df_raw = pd.read_excel(file_path, sheet_name=target_sheet, header=None, engine="openpyxl")
                header_row_idx = None
                for r in range(min(15, len(df_raw))):
                    row_vals = [str(v).strip() for v in df_raw.iloc[r].tolist()[:15]]
                    if any(x in row_vals for x in ["Input IPSS", "Input PSS", "Total WIP", "E3157 (Develop)"]):
                        header_row_idx = r; break
                
                if header_row_idx is not None:
                    df = pd.read_excel(file_path, sheet_name=target_sheet, header=header_row_idx, engine="openpyxl")
                    # Clean headers
                    df.columns = [str(c).replace("\n"," ").strip() for c in df.columns]
                    df.columns = [" ".join(c.split()) for c in df.columns]
                    # Keep valid date rows
                    df = df[pd.to_datetime(df.iloc[:, 0], errors="coerce").notna()].copy()
                    df.iloc[:, 0] = pd.to_datetime(df.iloc[:, 0], errors="coerce")
                    
                    def _clean(v):
                        if pd.isna(v) or v is None: return 0
                        if isinstance(v, (int, float)): return v
                        try: return float(str(v).replace(",","").strip().replace("%",""))
                        except: return 0
                    for col in df.columns[1:]:
                        if col != "Date" and col != "Ngày":
                            df[col] = df[col].apply(_clean)
                    return df
        except Exception as e:
            print(f"Error reading summary with pandas: {e}")

    # Priority 2: Fallback to openpyxl object (wb_instance)
    if wb_instance:
        try:
            target_sheet = next((name for name in wb_instance.sheetnames if name.strip().upper() == "SUMMARY"), None)
            if target_sheet:
                ws = wb_instance[target_sheet]
                # Manual fallback logic if needed...
        except: pass
    
    return pd.DataFrame()


def summary_to_trend_df(df_summary: pd.DataFrame) -> pd.DataFrame:
    """
    Convert raw SUMMARY sheet DataFrame sang format internal.
    Uses robust column name matching and index fallback for steps.
    """
    if df_summary.empty:
        return pd.DataFrame()

    df = df_summary.copy()
    raw_cols = list(df.columns)
    
    # Map names to internal keys (Fuzzy/Clean)
    name_map = {
        "Ngày": "Date", "Date": "Date",
        "Năm": "Year", "Year": "Year",
        "Tháng": "Month", "Month": "Month",
        "Tuần ISO": "Week", "ISO Week": "Week",
        "MONTH": "MONTH_Label",
        "Week": "Week_Label",
        "Input PSS": "Input_PSS", "Input_IPSS": "Input_IPSS",
        "EPI Shipment": "EPI_Shipment", "Total WIP": "Total_WIP",
        "Total Move": "Total_Move", "E3157 (Develop)": "E3157_Move",
        "PR RW (qty)": "PR_RW_Qty", "DF Rate %": "DF_Rate"
    }
    
    # Detect Layout by looking for key headers
    col_str = " ".join([str(c).replace("\n"," ") for c in raw_cols])
    is_v03 = ("Input IPSS" in col_str or "Tuần ISO" in col_str)

    rename_dict = {}
    used_indices = set()
    
    # 1. Map Fixed Metrics by Name
    for ci, col in enumerate(raw_cols):
        clean_name = str(col).replace("\n", " ").strip()
        if clean_name in name_map:
            rename_dict[col] = name_map[clean_name]
            used_indices.add(ci)

    # 2. Map Steps (WIP/Move batches)
    # Using Indices for steps is most reliable in V03 layout if names collide
    if is_v03:
        wip_start, move_start, pr_start = 14, 28, 42
    else:
        wip_start, move_start, pr_start = 9, 23, 37

    for i, step in enumerate(STEPS):
        w_idx, m_idx = wip_start + i, move_start + i
        if len(raw_cols) > w_idx:
            rename_dict[raw_cols[w_idx]] = f"WIP_{step}"
            used_indices.add(w_idx)
        if len(raw_cols) > m_idx:
            rename_dict[raw_cols[m_idx]] = f"Move_{step}"
            used_indices.add(m_idx)
            
    # 3. Map PR Columns
    for i, code in enumerate(PR_COLS):
        p_idx = pr_start + i
        if len(raw_cols) > p_idx:
            rename_dict[raw_cols[p_idx]] = f"PR_{code}"
            used_indices.add(p_idx)

    # 4. Standardize First column as Date if not mapped
    if 0 not in used_indices:
        rename_dict[raw_cols[0]] = "Date"
        used_indices.add(0)

    df = df.rename(columns=rename_dict)
    
    # 5. Cleanup and Fill Labels
    if "Date" not in df.columns: return pd.DataFrame()
    df["_dt"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df[df["_dt"].notna()].copy().reset_index(drop=True)
    df["Date"] = df["_dt"].dt.strftime("%Y-%m-%d")
    df["Label"] = df["Date"]

    # Fill Year/Month/Week if missing or 0
    # Use _si(v) on entries first to clean "2 026" etc.
    for col in ["Year", "Month", "Week"]:
        if col not in df.columns: df[col] = 0
        df[col] = df[col].apply(_si)
        mask = (df[col] == 0)
        if mask.any():
            if col == "Year":  df.loc[mask, "Year"] = df.loc[mask, "_dt"].dt.year
            if col == "Month": df.loc[mask, "Month"] = df.loc[mask, "_dt"].dt.month
            if col == "Week":  df.loc[mask, "Week"] = df.loc[mask, "_dt"].dt.isocalendar().week

    if "MONTH_Label" not in df.columns: df["MONTH_Label"] = ""
    ml_mask = df["MONTH_Label"].fillna("").astype(str).str.strip().isin(["","0","0.0","None"])
    if ml_mask.any():
        df.loc[ml_mask, "MONTH_Label"] = df.loc[ml_mask, "_dt"].dt.strftime("%Y-%m")

    if "Week_Label" not in df.columns: df["Week_Label"] = ""
    wl_mask = df["Week_Label"].fillna("").astype(str).str.strip().isin(["","0","0.0","None"])
    if wl_mask.any():
        from modules.calculator import get_week_label
        df.loc[wl_mask, "Week_Label"] = df["_dt"].apply(lambda x: get_week_label(x))

    # 6. Final Numerical Cleaning
    num_cols = ["Input_PSS","Input_IPSS","EPI_Shipment","Total_WIP","Total_Move","E3157_Move","PR_RW_Qty","DF_Rate"]
    for col in df.columns:
        if col.startswith(("WIP_","Move_","PR_")) or col in num_cols:
            df[col] = df[col].apply(_si)

    return df.drop(columns=["_dt"]).drop_duplicates(subset=["Date"], keep="last")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4: WRITE SUMMARY / WEEKLY / MONTHLY SHEETS
# ══════════════════════════════════════════════════════════════════════════════

def _write_data_sheet(wb, sheet_name: str, title: str, label_name: str,
                      rows_data: list, h_color: str = C_NAVY, date_fmt: bool = False):
    """
    Write SUMMARY, WEEKLY, or MONTHLY sheet with V03 column layout:
    Label | Input PSS/IPSS | EPI Ship | Total WIP/Move | E3157 | PR RW | DF Rate%
         | WIP×14 steps | Move×14 steps | PR01..PR33,PR99
    """
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
    ws.sheet_view.showGridLines = False

    # Define Time columns based on sheet - matching Apr 15 Golden File
    # ── Cấu trúc cột ──────────────────────────────────────────────────────────
    if sheet_name == "WEEKLY":
        time_hdrs = ["Year", "Week"]
        time_keys = ["Year", "Week"]
        sum_label_col = "$F"
    elif sheet_name == "MONTHLY":
        time_hdrs = ["Year", "Month"]
        time_keys = ["Year", "Month"]
        sum_label_col = "$E"
    else:
        time_hdrs = ["Year", "Month", "ISO Week", "MONTH", "Week"]
        time_keys = ["Year", "Month", "Week", "MONTH_Label", "Week_Label"]
        sum_label_col = None

    t_cnt = len(time_hdrs)
    total_cols = 1 + t_cnt + 8 + (len(STEPS) * 2) + len(PR_COLS)

    # ── Row 1: Title (merged) ─────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    cl = ws.cell(1, 1, title)
    cl.fill = _fill(h_color); cl.font = _font(True, "FFFFFF", 12)
    cl.alignment = _align("center"); cl.border = BT
    ws.row_dimensions[1].height = 28

    # ── Row 2: Column headers ─────────────────────────────────────────────────
    base_hdrs = [label_name] + time_hdrs + [
        "Input PSS", "Input IPSS", "EPI Shipment",
        "Total WIP", "Total Move", "E3157 (Develop)", "PR RW (qty)", "DF Rate %"
    ]
    hdrs = base_hdrs + list(STEPS) + list(STEPS) + list(PR_COLS)

    bg2 = ([h_color] * (1 + t_cnt) +
           [C_NAVY] * 3 +
           [C_NAVY, C_NAVY, C_TEAL, C_PURP, C_PURP] +
           [C_TEAL if s == DEVELOP_STEP else C_BLUE for s in STEPS] +
           [C_TEAL if s == DEVELOP_STEP else C_BLUE for s in STEPS] +
           [C_PURP] * len(PR_COLS))

    for ci, h in enumerate(hdrs, 1):
        _hc(ws, 2, ci, h, bg=bg2[ci - 1], s=9, wrap=True)
    ws.row_dimensions[2].height = 35

    # ── Vị trí cột trong sheet này ────────────────────────────────────────────
    m_start  = 2 + t_cnt
    w_start  = m_start + 8
    mv_start = w_start + len(STEPS)
    pr_start = mv_start + len(STEPS)

    wip_start_col  = get_column_letter(w_start)
    wip_end_col    = get_column_letter(w_start + len(STEPS) - 1)
    move_start_col = get_column_letter(mv_start)
    move_end_col   = get_column_letter(mv_start + len(STEPS) - 1)
    pr_sum_start   = get_column_letter(pr_start)
    pr_sum_end     = get_column_letter(pr_start + len(PR_COLS) - 1)

    SUM_OFFSET = 3
    SUM_DATA_ROW = 3

    # ── Data rows (bắt đầu từ row 3) ─────────────────────────────────────────
    for ri, row_dict in enumerate(rows_data, SUM_DATA_ROW):
        bg = "FFFFFF" if ri % 2 == 0 else "F8F9FA"
        lbl = row_dict.get("Label", row_dict.get("Date", ""))
        has_data = bool(row_dict.get("Total_Move", 0) or row_dict.get("PR_RW_Qty", 0)
                        or row_dict.get("E3157_Move", 0))

        if date_fmt:
            cl = ws.cell(ri, 1, _date(lbl))
            cl.fill = _fill(bg); cl.font = _font(b=has_data)
            cl.alignment = _align("center"); cl.border = BT; cl.number_format = DATE_FMT
        else:
            _dc(ws, ri, 1, lbl, bg=bg, h="center", b=has_data)

        # Time columns — dùng raw value, không _si() để tránh string -> 0
        for ti, key in enumerate(time_keys, 2):
            raw_val = row_dict.get(key)
            if sheet_name == "SUMMARY" and key == "MONTH_Label":
                # Formula giong file mau: =+B3&"-"&IF(C3<10,"0"&C3,C3)
                f_mo = f'=+B{ri}&"-"&IF(C{ri}<10,"0"&C{ri},C{ri})'
                cl = ws.cell(ri, ti, f_mo)
                cl.fill = _fill(bg); cl.font = _font()
                cl.alignment = _align("center"); cl.border = BT
            elif isinstance(raw_val, str):
                _dc(ws, ri, ti, raw_val, bg=bg, h="center")
            else:
                _dc(ws, ri, ti, _si(raw_val), bg=bg)

        # Column refs
        e3157_col  = get_column_letter(m_start + 5)
        pr_rw_col  = get_column_letter(m_start + 6)

        # SUM_OFFSET: diff between SUMMARY t_cnt(5) and this sheet t_cnt(2)
        SUM_OFFSET   = 5 - len(time_keys)   # 3 for WEEKLY/MONTHLY, 0 for SUMMARY
        SUM_DATA_ROW = 3
        s_lbl = ("$F" if sheet_name == "WEEKLY" else
                 "$E" if sheet_name == "MONTHLY" else None)

        # SUMMARY formula ranges
        wip_start_col  = get_column_letter(w_start)
        wip_end_col    = get_column_letter(w_start + len(STEPS) - 1)
        move_start_col = get_column_letter(mv_start)
        move_end_col   = get_column_letter(mv_start + len(STEPS) - 1)
        pr_s = get_column_letter(pr_start)
        pr_e = get_column_letter(pr_start + len(PR_COLS) - 1)

        # Metric columns (Input PSS → PR RW)
        for ci, key in enumerate(["Input_PSS","Input_IPSS","EPI_Shipment",
                                    "Total_WIP","Total_Move","E3157_Move","PR_RW_Qty"], m_start):
            v    = _si(row_dict.get(key, 0))
            bg_c = "E3F2FD" if key == "E3157_Move" and v > 0 else bg
            s_ci  = ci + SUM_OFFSET
            s_col = get_column_letter(s_ci)

            if sheet_name == "SUMMARY":
                if key == "Total_WIP":
                    cl = ws.cell(ri, ci, f"=+SUM({wip_start_col}{ri}:{wip_end_col}{ri})")
                    cl.fill = _fill(bg_c); cl.border = BT; cl.number_format = '#,##0'
                elif key == "Total_Move":
                    cl = ws.cell(ri, ci, f"=+SUM({move_start_col}{ri}:{move_end_col}{ri})")
                    cl.fill = _fill(bg_c); cl.border = BT; cl.number_format = '#,##0'
                elif key == "PR_RW_Qty":
                    cl = ws.cell(ri, ci, f"=+SUM({pr_s}{ri}:{pr_e}{ri})")
                    cl.fill = _fill(bg_c); cl.border = BT; cl.number_format = '#,##0'
                else:
                    _dc(ws, ri, ci, v, bg=bg_c, b=(key == "E3157_Move" and v > 0))
            else:
                # WEEKLY / MONTHLY — cong thuc giong file mau ngay 15
                if key == "Total_WIP":
                    # LOOKUP lay WIP ngay cuoi cua ky — J = Total WIP trong SUMMARY
                    sum_j = get_column_letter(10)
                    f = (f"=+LOOKUP(2,1/(SUMMARY!{s_lbl}${SUM_DATA_ROW}:"
                         f"{s_lbl}$1048576=A{ri}),"
                         f"SUMMARY!{sum_j}${SUM_DATA_ROW}:{sum_j}$1048576)")
                    cl = ws.cell(ri, ci, f)
                    cl.fill = _fill(bg_c); cl.border = BT; cl.number_format = '#,##0'
                else:
                    # SUMIFS cong tong theo tuan/thang
                    f = (f"=+SUMIFS(SUMMARY!{s_col}${SUM_DATA_ROW}:{s_col}$1048576,"
                         f"SUMMARY!{s_lbl}${SUM_DATA_ROW}:{s_lbl}$1048576,"
                         f"{sheet_name}!$A{ri})")
                    cl = ws.cell(ri, ci, f)
                    cl.fill = _fill(bg_c); cl.border = BT; cl.number_format = '#,##0'
                    if key == "E3157_Move": cl.font = _font(b=True)

        # DF Rate: =+IFERROR(PR_RW/E3157,"") format 0.00% (Excel auto *100)
        f_df  = f'=+IFERROR({pr_rw_col}{ri}/{e3157_col}{ri},"")'
        cl_df = ws.cell(ri, m_start + 7, f_df)
        dfr_val = float(row_dict.get("DF_Rate", 0) or 0)
        cl_df.fill = _fill("FFCDD2" if dfr_val > 15 else "FFF9C4" if dfr_val > 10 else bg)
        cl_df.number_format = '0.00%'; cl_df.border = BT

        # WIP Steps — LOOKUP (ngay cuoi ky) cho WEEKLY/MONTHLY
        for i, s in enumerate(STEPS):
            ci   = w_start + i
            v    = _si(row_dict.get(f"WIP_{s}", 0))
            bg_c = "E3F2FD" if s == DEVELOP_STEP and v > 0 else bg
            if sheet_name == "SUMMARY":
                _dc(ws, ri, ci, v, bg=bg_c)
            else:
                sum_wip_col = get_column_letter(w_start + SUM_OFFSET + i)
                f = (f"=+LOOKUP(2,1/(SUMMARY!{s_lbl}${SUM_DATA_ROW}:"
                     f"{s_lbl}$1048576=A{ri}),"
                     f"SUMMARY!{sum_wip_col}${SUM_DATA_ROW}:{sum_wip_col}$1048576)")
                cl = ws.cell(ri, ci, f)
                cl.fill = _fill(bg_c); cl.border = BT; cl.number_format = '#,##0'

        # Move Steps — SUMIFS cho WEEKLY/MONTHLY
        for i, s in enumerate(STEPS):
            ci   = mv_start + i
            v    = _si(row_dict.get(f"Move_{s}", 0))
            bg_c = "B3E5FC" if s == DEVELOP_STEP and v > 0 else bg
            if sheet_name == "SUMMARY":
                _dc(ws, ri, ci, v, bg=bg_c)
            else:
                sum_mv_col = get_column_letter(mv_start + SUM_OFFSET + i)
                f = (f"=+SUMIFS(SUMMARY!{sum_mv_col}${SUM_DATA_ROW}:{sum_mv_col}$1048576,"
                     f"SUMMARY!{s_lbl}${SUM_DATA_ROW}:{s_lbl}$1048576,"
                     f"{sheet_name}!$A{ri})")
                cl = ws.cell(ri, ci, f)
                cl.fill = _fill(bg_c); cl.border = BT; cl.number_format = '#,##0'

        # PR Codes — SUMIFS cho WEEKLY/MONTHLY
        for i, c in enumerate(PR_COLS):
            ci = pr_start + i
            v  = _si(row_dict.get(f"PR_{c}", 0))
            if sheet_name == "SUMMARY":
                _dc(ws, ri, ci, v, bg="FFE5E5" if v > 0 else bg)
            else:
                sum_pr_col = get_column_letter(pr_start + SUM_OFFSET + i)
                f = (f"=+SUMIFS(SUMMARY!{sum_pr_col}${SUM_DATA_ROW}:{sum_pr_col}$1048576,"
                     f"SUMMARY!{s_lbl}${SUM_DATA_ROW}:{s_lbl}$1048576,"
                     f"{sheet_name}!$A{ri})")
                cl = ws.cell(ri, ci, f)
                cl.fill = _fill("FFE5E5" if v > 0 else bg); cl.border = BT
                cl.number_format = '#,##0'

    # Column widths
    ws.column_dimensions["A"].width = 14
    for ci in range(2, m_start + 9):
        ws.column_dimensions[get_column_letter(ci)].width = 11
    for ci in range(m_start + 9, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 8

    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{get_column_letter(total_cols)}2"
    return ws

def _write_pr_rw_lot_list(wb, df_new_hold: pd.DataFrame):
    """Robust V03 Update: Scans headers for exact column mapping and ensures sorting."""
    sheet_name = "PR RW LOT LIST"
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]

    # 1. Map Header Strings to Column Indices (1-based)
    h_map = {}
    for c in range(1, 40):
        v = str(ws.cell(1, c).value or "").strip().upper()
        if not v: continue
        if "YEAR" in v: h_map["YEAR"] = c
        elif "MONTH" in v and len(v) < 6: h_map["MONTH"] = c
        elif "WEEK" in v: h_map["WEEK"] = c
        elif "DAY" == v or "DATE" == v: h_map["DAY"] = c
        elif "PROCESS" == v: h_map["Process"] = c
        elif "STEP" == v: h_map["Step"] = c
        elif "DESCRIPTION" in v: h_map["Description"] = c
        elif "PRODUCT GROUP" in v: h_map["Product Group"] = c
        elif "PART ID" in v: h_map["Part ID"] = c
        elif "LOT" == v: h_map["Lot"] = c
        elif "UNIT" == v: h_map["Unit"] = c
        elif "QTY" == v or "QUANTITY" == v: h_map["Qty"] = c
        elif "HOLD TIME" in v: h_map["Hold Time(Hour)"] = c
        elif "HOLD DATE" in v: h_map["Hold date and time"] = c
        elif "HOLD OPERATOR" in v: h_map["Hold Operator"] = c
        elif "HOLD REASON" in v: h_map["Hold reason"] = c
        elif "RELEASE TIME" in v: h_map["Release time"] = c
        elif "RELEASE OPERATOR" in v: h_map["Release Operator"] = c
        elif "RELEASE REASON" in v: h_map["Release reason"] = c
        elif "HOLD CODE" in v: h_map["Hold Code"] = c
        elif "CODE TYPE" in v: h_map["Code Type"] = c
        elif "PRIORITY" in v: h_map["Priority"] = c

    # 2. Extract RAW data flexibly
    def get_raw_val(row, keywords, default=""):
        for k in keywords:
            for c in df_new_hold.columns:
                if str(c).strip().upper() == k.upper(): return row[c] if pd.notna(row[c]) else default
        for k in keywords:
            for c in df_new_hold.columns:
                if k.upper() in str(c).strip().upper(): return row[c] if pd.notna(row[c]) else default
        return default

    # 3. Read existing sheet data
    data_list = []
    if ws.max_row >= 2:
        for r in range(2, ws.max_row + 1):
            item = {}
            for key, col_idx in h_map.items():
                item[key] = ws.cell(r, col_idx).value
            if item.get("Lot"): data_list.append(item)

    # 4. Process NEW data from RAW
    if not df_new_hold.empty:
        h_col = next((c for c in df_new_hold.columns if "HOLD" in str(c).upper() and "DATE" in str(c).upper()), None)
        l_col = next((c for c in df_new_hold.columns if "LOT" == str(c).strip().upper()), None)
        c_col = next((c for c in df_new_hold.columns if "HOLD" in str(c).upper() and "CODE" in str(c).upper()), None)
        pg_col = next((c for c in df_new_hold.columns if "PRODUCT" in str(c).upper() and "GROUP" in str(c).upper()), None)

        if h_col and l_col and c_col:
            df_pr = df_new_hold.copy()
            if pg_col: df_pr = df_pr[df_pr[pg_col].astype(str).str.upper().str.contains("IPSS", na=False)]
            df_pr = df_pr[df_pr[c_col].astype(str).str.upper().str.startswith("PR")]

            for _, row in df_pr.iterrows():
                h_dt = pd.to_datetime(row[h_col], errors="coerce")
                lot = str(row[l_col]).strip(); code = str(row[c_col]).strip()
                if not lot or pd.isna(h_dt): continue
                
                h_dt_str = h_dt.strftime("%Y-%m-%d %H:%M:%S")
                # Deduplicate by Lot + Hold DateTime + Code
                existing = None
                for x in data_list:
                    ex_dt = x.get("Hold date and time")
                    ex_dt_str = ex_dt.strftime("%Y-%m-%d %H:%M:%S") if isinstance(ex_dt, datetime) else str(ex_dt)
                    if x.get("Lot") == lot and x.get("Hold Code") == code and ex_dt_str == h_dt_str:
                        existing = x; break
                
                rel_val = get_raw_val(row, ["Release time"])
                rel_dt = pd.to_datetime(rel_val, errors="coerce") if rel_val else pd.NaT

                if existing:
                    if not existing.get("Release time") and pd.notna(rel_dt):
                        existing["Release time"] = rel_dt
                        existing["Release Operator"] = get_raw_val(row, ["Release Operator"])
                        existing["Release reason"] = get_raw_val(row, ["Release reason"])
                else:
                    day_val = (h_dt + pd.Timedelta(hours=4)).normalize()
                    data_list.append({
                        "YEAR": day_val.year, "MONTH": day_val.month, "WEEK": int(day_val.isocalendar().week), "DAY": day_val,
                        "Process": get_raw_val(row, ["Process"]), "Step": get_raw_val(row, ["Step"]),
                        "Description": get_raw_val(row, ["Description", "Desc"]), "Product Group": get_raw_val(row, ["Product Group"]),
                        "Part ID": get_raw_val(row, ["Part ID"]), "Lot": lot, "Unit": get_raw_val(row, ["Unit"]),
                        "Qty": _si(get_raw_val(row, ["Qty"], 0)), "Hold Time(Hour)": get_raw_val(row, ["Hold Time"]),
                        "Hold date and time": h_dt, "Hold Operator": get_raw_val(row, ["Hold Operator"]),
                        "Hold reason": get_raw_val(row, ["Hold reason"]), "Release time": rel_dt if pd.notna(rel_dt) else None,
                        "Release Operator": get_raw_val(row, ["Release Operator"]), "Release reason": get_raw_val(row, ["Release reason"]),
                        "Hold Code": code, "Code Type": get_raw_val(row, ["Code Type"]), "Priority": get_raw_val(row, ["Priority"])
                    })

    # 5. SORT: Chronological (Old -> New)
    def ms_key(x):
        d = x.get("Hold date and time")
        if isinstance(d, datetime): return d
        try: return pd.to_datetime(d, errors="coerce")
        except: return datetime(1900, 1, 1)

    data_list.sort(key=lambda x: ms_key(x) or datetime(1900, 1, 1))

    # 6. REWRITE Sheet
    ws.delete_rows(2, ws.max_row + 1)
    y_fill = _fill("FFFF00")
    for ri, item in enumerate(data_list, 2):
        for key, col_idx in h_map.items():
            cell = ws.cell(ri, col_idx)
            val = item.get(key)
            if str(val).lower() in ["nan", "none"]: val = ""
            cell.value = val
            if key == "DAY": cell.number_format = DATE_FMT
            elif key in ["Hold date and time", "Release time"]: cell.number_format = DT_FMT
            cell.border = BT; cell.font = _font(s=9)
            if not item.get("Release time"): cell.fill = y_fill
            else: cell.fill = PatternFill(fill_type=None)
    
    ws.freeze_panes = "A2"
    return ws




# ══════════════════════════════════════════════════════════════════════════════
# HELPER: FREEZE OLD FORMULA ROWS → VALUE (tăng tốc độ mở file Excel)
# ══════════════════════════════════════════════════════════════════════════════

def _freeze_summary_rows(ws_sum, ws_val_sum, keep_rows: int = 10, header_row: int = 2):
    """
    Freeze SUMMARY: chuyển công thức → giá trị tĩnh cho tất cả dòng cũ.

    Phân loại 2 nhóm công thức trong SUMMARY:
      ① Tính được trực tiếp (không cần wb_val):
         Col  5 (E): MONTH label  → "YYYY-MM" từ Year(B) + Month(C)
         Col 10 (J): Total WIP    → SUM WIP step cols (O=15..AB=28), vì WIP steps là plain values
      ② Cần wb_val (data_only) vì là SUMIFS tham chiếu sheet khác:
         Col 43..76  : PR01..PR99  → =SUMIFS('PR RW LOT LIST'!...)
         Col 13 (M)  : PR RW qty   → =SUM(AQ:BX) của các SUMIFS ở trên
         Col 14 (N)  : DF Rate %   → sau khi có col 13, tính từ col 13 / col 12
    """
    n_steps = len(STEPS)    # 14
    n_pr    = len(PR_COLS)  # 34

    COL_YEAR      = 2;  COL_MONTH_NUM = 3
    COL_MO_LBL   = 5   # MONTH  =+B&"-"&IF(C<10,"0"&C,C)
    COL_TOTAL_WIP = 10  # =SUM(O:AB)    — WIP steps là plain values
    COL_E3157     = 12  # plain value
    COL_PR_RW     = 13  # =SUM(AQ:BX)  — AQ:BX là SUMIFS
    COL_DF_RATE   = 14  # =IFERROR(M/L,"")
    WIP_START     = 15  # col O
    PR_START      = 43  # col AQ

    def _is_f(cell):
        return (cell.data_type == 'f' or
                (isinstance(cell.value, str) and cell.value.startswith('=')))

    def _val_or_none(r, c):
        """Lấy giá trị đã tính từ wb_val; trả về None nếu không có."""
        if ws_val_sum is None:
            return None
        v = ws_val_sum.cell(r, c).value
        return v  # có thể là None nếu chưa qua Excel

    data_rows = []
    for r in range(header_row + 1, ws_sum.max_row + 1):
        v = ws_sum.cell(r, 1).value
        if v is not None and str(v).strip() not in ("", "None"):
            data_rows.append(r)

    if len(data_rows) <= keep_rows:
        return

    for r in data_rows[:-keep_rows]:
        # ── ① MONTH label: tính trực tiếp (Year & Month là plain values) ────
        c5 = ws_sum.cell(r, COL_MO_LBL)
        if _is_f(c5):
            yr = _si(ws_sum.cell(r, COL_YEAR).value)
            mo = _si(ws_sum.cell(r, COL_MONTH_NUM).value)
            c5.value = f"{yr}-{mo:02d}" if yr > 0 and mo > 0 else ""

        # ── ① Total WIP: tính trực tiếp (WIP step cols là plain values) ─────
        c10 = ws_sum.cell(r, COL_TOTAL_WIP)
        if _is_f(c10):
            c10.value = sum(_si(ws_sum.cell(r, WIP_START + i).value)
                            for i in range(n_steps))

        # ── ② PR code cols 43..76: SUMIFS → lấy từ wb_val ───────────────────
        # Track xem bao nhiêu PR code cols đã được freeze thành plain value
        pr_frozen_count = 0
        for i in range(n_pr):
            cp = ws_sum.cell(r, PR_START + i)
            if _is_f(cp):
                computed = _val_or_none(r, PR_START + i)
                if computed is not None:
                    cp.value = computed
                    pr_frozen_count += 1
                # Nếu None (chưa qua Excel) → giữ SUMIFS công thức, an toàn
            else:
                # Đã là plain value → đã freeze rồi
                pr_frozen_count += 1

        # ── ② PR RW qty (col 13): lấy từ wb_val, fallback tổng PR cols ──────
        c13 = ws_sum.cell(r, COL_PR_RW)
        if _is_f(c13):
            computed = _val_or_none(r, COL_PR_RW)
            if computed is not None:
                c13.value = computed
                pr_sum = _si(computed)
            elif pr_frozen_count == n_pr:
                # Tất cả PR code cols đã là plain values → fallback sum an toàn
                pr_sum = sum(_si(ws_sum.cell(r, PR_START + i).value)
                             for i in range(n_pr))
                c13.value = pr_sum
            else:
                # Vẫn còn PR code cols dạng SUMIFS formula chưa được tính
                # → KHÔNG ghi đè formula PR_RW bằng 0, giữ nguyên formula
                pr_sum = 0
                # c13.value được giữ nguyên (là formula =SUM(AQ:BX))
        else:
            pr_sum = _si(c13.value)

        # ── ② DF Rate (col 14): tính từ pr_sum / E3157 ───────────────────────
        c14 = ws_sum.cell(r, COL_DF_RATE)
        if _is_f(c14):
            e3157 = _si(ws_sum.cell(r, COL_E3157).value)
            # Chỉ freeze DF Rate khi đã có pr_sum thực (không phải 0 do chưa tính được)
            if pr_frozen_count == n_pr or pr_sum > 0:
                c14.value = round(pr_sum / e3157, 6) if e3157 > 0 else 0.0
            # else: giữ nguyên formula DF Rate


def _freeze_old_rows(ws_write, ws_read, header_row: int = 2, keep_rows: int = 10):
    """
    Freeze WEEKLY / MONTHLY: chuyển SUMIFS/LOOKUP → giá trị tĩnh cho dòng cũ.
    Dùng ws_read (data_only) để lấy giá trị Excel đã tính sẵn.
    Nếu giá trị là None (file chưa qua Excel) → giữ nguyên công thức (an toàn).
    """
    data_rows = []
    for r in range(header_row + 1, ws_write.max_row + 1):
        v = ws_write.cell(r, 1).value
        if v is not None and str(v).strip() not in ("", "None"):
            data_rows.append(r)

    if len(data_rows) <= keep_rows:
        return

    rows_to_freeze = data_rows[:-keep_rows]
    max_col = ws_write.max_column

    for r in rows_to_freeze:
        for c in range(1, max_col + 1):
            cell_w = ws_write.cell(r, c)
            is_formula = (
                cell_w.data_type == 'f' or
                (isinstance(cell_w.value, str) and cell_w.value.startswith('='))
            )
            if not is_formula:
                continue
            computed = ws_read.cell(r, c).value
            if computed is None:
                continue   # Chưa có giá trị tính → giữ công thức
            cell_w.value = computed


# ══════════════════════════════════════════════════════════════════════════════
# MAIN UPDATE FUNCTION
# ══════════════════════════════════════════════════════════════════════════════

def update_report_v3(
    template_path: str,
    output_path: str,
    raw_data: dict,
    today_dt: datetime,
    week_start_day: int = 4,
    epi_shipment: int = 0,
) -> str:
    """
    Refined Update Logic:
    1. If output exists, load it. Else copy template.
    2. Update RAW sheets (Write over existing).
    3. Calculate data for Today + Yesterday.
    4. FIND the rows in SUMMARY and update only the metric cells.
    5. Preserve all existing formulas/formatting.
    """
    # 1. Setup File
    if not os.path.exists(output_path):
        if template_path and os.path.exists(template_path):
            shutil.copy2(template_path, output_path)
        else:
            raise FileNotFoundError("Template not found and output doesn't exist.")

    wb = openpyxl.load_workbook(output_path)
    # ── Mẹo: Load data_only để đọc giá trị ngày nếu dùng công thức ──────────────────
    wb_val = openpyxl.load_workbook(output_path, data_only=True)
    ws_val = wb_val["SUMMARY"] if "SUMMARY" in wb_val.sheetnames else None

    # 2. RAW Sheets Update
    for sheet_name, key in [
        ("RAW_HOLD HISTORY", "hold_history"),
        ("RAW_WIP",          "wip"),
        ("RAW_MOVE",         "move"),
        ("RAW_INPUT",        "input"),
    ]:
        if sheet_name in wb.sheetnames:
            _write_raw_sheet(wb, sheet_name, raw_data.get(key, pd.DataFrame()))

    # 3. Update SUMMARY sheet
    ws_sum = wb["SUMMARY"] if "SUMMARY" in wb.sheetnames else None
    if ws_sum and ws_val:
        # Vietnamese -> English header normalization map (standardize to English)
        VN_TO_EN = {
            "Ngày": "Date", "Năm": "Year", "Tháng": "Month",
            "Tuần ISO": "ISO Week", "Tuần": "Week",
        }
        # Normalize header row in-place (one-time migration to English)
        header_row = 2
        for c in range(1, ws_sum.max_column + 1):
            raw_val = str(ws_sum.cell(header_row, c).value or "").strip()
            if raw_val in VN_TO_EN:
                ws_sum.cell(header_row, c).value = VN_TO_EN[raw_val]

        # Build header_map from the (now-English) header row
        header_map = {}
        for c in range(1, ws_sum.max_column + 1):
            val = str(ws_sum.cell(header_row, c).value or "").strip()
            if val: header_map[val] = c

        # Fallback: if Date not in row 2, try row 1
        if "Date" not in header_map and "Day" not in header_map:
            header_row = 1
            header_map = {}
            for c in range(1, ws_sum.max_column + 1):
                raw_val = str(ws_sum.cell(header_row, c).value or "").strip()
                en_val = VN_TO_EN.get(raw_val, raw_val)
                if raw_val in VN_TO_EN:
                    ws_sum.cell(header_row, c).value = en_val
                if en_val: header_map[en_val] = c
            
        # Last fallback: check row 1 if 'Date' still not found
        if "Date" not in header_map:
            header_row = 1
            header_map = {}
            for c in range(1, ws_sum.max_column + 1):
                raw_val = str(ws_sum.cell(header_row, c).value or "").strip()
                en_val = VN_TO_EN.get(raw_val, raw_val)
                if raw_val in VN_TO_EN:
                    ws_sum.cell(header_row, c).value = en_val
                if en_val: header_map[en_val] = c

        target_norm = pd.Timestamp(today_dt).normalize()
        yesterday_norm = target_norm - pd.Timedelta(days=1)
        update_dates = [target_norm, yesterday_norm]

        for d_ts in update_dates:
            dt_obj = d_ts.to_pydatetime()
            row_data = calc_daily_row_from_raw(raw_data, dt_obj, week_start_day)
            if d_ts == target_norm and epi_shipment > 0:
                row_data["EPI_Shipment"] = epi_shipment
            
            # Find row index (Sử dụng ws_val để đọc kết quả công thức)
            target_row_idx = None
            for r in range(header_row + 1, ws_val.max_row + 200):
                cell_val = ws_val.cell(r, 1).value
                if cell_val:
                    try:
                        if isinstance(cell_val, (datetime, pd.Timestamp)):
                            row_dt = pd.Timestamp(cell_val).normalize()
                        else:
                            row_dt = pd.to_datetime(str(cell_val).strip(), errors="coerce")
                            if row_dt: row_dt = row_dt.normalize()
                        
                        if row_dt == d_ts:
                            target_row_idx = r; break
                    except: pass
            
            # 4. Write data to row
            if not target_row_idx:
                for r in range(header_row + 1, 10000):
                    if not ws_val.cell(r, 1).value:
                        target_row_idx = r; break
                ws_sum.cell(target_row_idx, 1).value = dt_obj
                ws_sum.cell(target_row_idx, 1).number_format = "yyyy-mm-dd"
            
            # Inheritance: Copy format from row above (if it exists)
            if target_row_idx > (header_row + 1):
                _copy_row_format(ws_sum, target_row_idx - 1, target_row_idx)

            # Map internal key -> column index (header_map is now all-English)
            key_map = {
                "Year":         header_map.get("Year"),
                "Month":        header_map.get("Month"),
                "Week":         header_map.get("ISO Week") or header_map.get("Week"),
                "MONTH_Label":  header_map.get("MONTH"),
                "Week_Label":   header_map.get("Week"),
                "Input_PSS":    header_map.get("Input PSS"),
                "Input_IPSS":   header_map.get("Input IPSS"),
                "EPI_Shipment": header_map.get("EPI Shipment"),
                "Total_WIP":    header_map.get("Total WIP"),
                "Total_Move":   header_map.get("Total Move"),
                "E3157_Move":   header_map.get("E3157 (Develop)"),
                "PR_RW_Qty":    header_map.get("PR RW (qty)"),
                "DF_Rate":      header_map.get("DF Rate %"),
            }
            # V03 fixed column indices (1-based): WIP O(15)-AB(28), Move AC(29)-AP(42), PR AQ(43)+
            w_start, m_start, p_start = 15, 29, 43

            # Write base metrics (non-step data)
            for k, val in row_data.items():
                if k.startswith(("WIP_", "Move_")) or (k.startswith("PR_") and k != "PR_RW_Qty"):
                    continue
                
                col_idx = key_map.get(k)
                if col_idx:
                    # PRESERVE FORMULAS: Check if existing cell in template has a formula
                    curr_cell = ws_sum.cell(target_row_idx, col_idx)
                    if curr_cell.data_type == 'f' or (isinstance(curr_cell.value, str) and curr_cell.value.startswith('=')):
                        continue # Do not overwrite Excel formulas
                        
                    v_new = _si(val)
                    v_old = _si(ws_val.cell(target_row_idx, col_idx).value)
                    
                    if d_ts == target_norm or v_new != 0 or v_old == 0:
                        ws_sum.cell(target_row_idx, col_idx).value = val
                        if k == "DF_Rate":
                            ws_sum.cell(target_row_idx, col_idx).number_format = '0.00"%"'

            # Write WIP / Move / PR steps by fixed column index
            for i, s in enumerate(STEPS):
                wv = row_data.get(f"WIP_{s}"); mv = row_data.get(f"Move_{s}")
                
                # Check for formulas in WIP/Move columns
                w_cell = ws_sum.cell(target_row_idx, w_start + i)
                m_cell = ws_sum.cell(target_row_idx, m_start + i)
                
                # RULE: ONLY write WIP for the target date (today). Never overwrite historical WIP snapshots.
                if d_ts == target_norm:
                    if wv is not None and w_cell.data_type != 'f' and not (isinstance(w_cell.value, str) and w_cell.value.startswith('=')):
                        w_cell.value = _si(wv)
                
                # Move values can be updated for historical dates as they are cumulative
                if mv is not None and m_cell.data_type != 'f' and not (isinstance(m_cell.value, str) and m_cell.value.startswith('=')):
                    m_cell.value = _si(mv)

                    
            for i, c in enumerate(PR_COLS):
                pv = row_data.get(f"PR_{c}")
                p_cell = ws_sum.cell(target_row_idx, p_start + i)
                if pv is not None and p_cell.data_type != 'f' and not (isinstance(p_cell.value, str) and p_cell.value.startswith('=')):
                    p_cell.value = _si(pv)


    # 4. Freeze old rows: công thức → giá trị tĩnh (chỉ giữ 10 dòng cuối có hàm)
    #    SUMMARY: tính trực tiếp từ các ô → không cần wb_val, luôn đúng
    #    WEEKLY/MONTHLY: dùng wb_val (cần file đã qua Excel ít nhất 1 lần)
    if "SUMMARY" in wb.sheetnames and "SUMMARY" in wb_val.sheetnames:
        _freeze_summary_rows(wb["SUMMARY"], wb_val["SUMMARY"], keep_rows=10, header_row=2)
    for _sname in ("WEEKLY", "MONTHLY"):
        if _sname in wb.sheetnames and _sname in wb_val.sheetnames:
            _freeze_old_rows(wb[_sname], wb_val[_sname], header_row=2, keep_rows=10)

    # 5. PR RW LOT LIST
    _write_pr_rw_lot_list(wb, raw_data.get("hold_history", pd.DataFrame()))

    # 6. Sheet Order & Save
    desired_order = ["KPI_TODAY","SUMMARY","WEEKLY","MONTHLY","PR RW LOT LIST",
                     "RAW_HOLD HISTORY","RAW_WIP","RAW_MOVE","RAW_INPUT","NOTE"]
    current = wb.sheetnames
    for sn in desired_order:
        if sn in current:
            idx = desired_order.index(sn)
            wb.move_sheet(sn, offset=idx - current.index(sn))

    wb.save(output_path)
    return output_path
